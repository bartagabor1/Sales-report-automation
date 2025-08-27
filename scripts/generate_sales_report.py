import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. Load data
data_path = "data/cleaned_sales_data.csv"
df = pd.read_csv(data_path)

# 2. Clean data: drop unnecessary column (unit_price_x)
df = df.drop(columns=["unit_price_x"])

# 3. Create pivot table: Revenue by country
pivot_country = df.pivot_table(
    index="country",
    values="revenue",
    aggfunc="sum"
).sort_values(by="revenue", ascending=False)

# 4. Create pivot table: Revenue by product category
pivot_category = df.pivot_table(
    index="category",
    values="revenue",
    aggfunc="sum"
).sort_values(by="revenue", ascending=False)

# 5. Create pivot table: Top 20 products by revenue
pivot_product = df.pivot_table(
    index="product_name",
    values="revenue",
    aggfunc="sum"
).sort_values(by="revenue", ascending=False).head(20)

# 6. Create Excel workbook
wb = Workbook()

# 7. Create main worksheet: Summary
ws_summary = wb.active
ws_summary.title = "Summary"

# Helper function to write DataFrame to worksheet with formatting
def write_df_to_ws(ws, df, start_row=1, start_col=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Format headers
            if r_idx == start_row or c_idx == start_col:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold color
    # Auto-adjust column widths
    for col_cells in ws.columns:
        max_length = 0
        col = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col].width = max_length + 2

# 8. Write pivot tables to worksheets
write_df_to_ws(ws_summary, pivot_country)
ws_category = wb.create_sheet(title="Revenue by Category")
write_df_to_ws(ws_category, pivot_category)
ws_product = wb.create_sheet(title="Top 20 Products")
write_df_to_ws(ws_product, pivot_product)

# 9. Save the workbook
output_path = "output/sales_report.xlsx"
wb.save(output_path)
print(f"✅ Report generated and saved to: {output_path}")
